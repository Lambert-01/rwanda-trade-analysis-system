#!/usr/bin/env python3
"""
Check Excel file sheets and identify commodity-related data
"""

import pandas as pd
import os
from openpyxl import load_workbook

def check_excel_sheets():
    excel_file = '../data/raw/2025Q1_Trade_report_annexTables.xlsx'

    if not os.path.exists(excel_file):
        print(f"Excel file not found: {excel_file}")
        return

    print("Loading Excel file...")
    wb = load_workbook(excel_file, read_only=True)

    print("\nAll available sheets:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"{i:2d}. {sheet_name}")

    print("\n" + "="*60)
    print("Looking for commodity-related sheets...")

    # Look for sheets that might contain commodity data
    commodity_keywords = ['commodity', 'export', 'import', 'reexport', 'product', 'good']
    potential_sheets = []

    for sheet_name in wb.sheetnames:
        sheet_lower = sheet_name.lower()
        if any(keyword in sheet_lower for keyword in commodity_keywords):
            potential_sheets.append(sheet_name)

    print(f"Potential commodity sheets: {potential_sheets}")

    # Check each potential sheet
    for sheet in potential_sheets:
        try:
            print(f"\n--- {sheet} ---")
            df = pd.read_excel(excel_file, sheet_name=sheet, nrows=10)  # First 10 rows
            print(f"Shape: {df.shape}")
            print(f"Columns: {list(df.columns)}")
            print("First few rows:")
            print(df.head(5))
            print("-" * 40)
        except Exception as e:
            print(f"Error reading {sheet}: {e}")

    wb.close()

if __name__ == "__main__":
    check_excel_sheets()